import os
import openpyxl
from flask import Flask, render_template_string, request

app = Flask(__name__)

# ==========================================
# CONSTANTES Y CONFIGURACIÓN
# ==========================================
FACTOR_ESCALA = 1000  # Factor para evitar problemas de coma regional en Excel
EXCEL_FILE = os.getenv("EXCEL_FILE", "Base_Nutricional.xlsx")


def buscar_comida_predeterminada(hoja_predeterminadas, nombre_ingresado):
    """
    Busca un alimento o plato en la hoja Comidas_Predeterminadas.
    NO utiliza atajos ni alias. Exige coincidencia exacta.
    
    Retorna un diccionario con los valores reales (divididos por 1000)
    o None si el comando/alimento no existe.
    """
    # Normalizamos el texto de búsqueda para evitar fallos por espacios o mayúsculas
    busqueda = str(nombre_ingresado).strip().lower()

    # Iteramos sobre las filas de la tabla (asumiendo encabezados en la fila 1)
    for fila in hoja_predeterminadas.iter_rows(min_row=2, values_only=True):
        codigo = fila[0]
        alimento = fila[1]

        if not alimento:
            continue

        # Comprobación de coincidencia exacta con la columna Alimento
        if str(alimento).strip().lower() == busqueda:
            return {
                "codigo": codigo,
                "alimento": alimento,
                "peso": (fila[2] or 0) / FACTOR_ESCALA,
                "calorias": (fila[3] or 0) / FACTOR_ESCALA,
                "proteinas": (fila[4] or 0) / FACTOR_ESCALA,
                "grasas": (fila[5] or 0) / FACTOR_ESCALA,
                "hidratos": (fila[6] or 0) / FACTOR_ESCALA,
                "fibras": (fila[7] or 0) / FACTOR_ESCALA,
            }

    # Si no coincide exactamente con nada, el programa desconoce el comando
    return None


def registrar_alimento_en_usuario(hoja_usuario, datos_alimento):
    """
    Guarda el alimento en la hoja del usuario multiplicando
    cada valor numérico por 1000 y convirtiéndolo a entero.
    """
    siguiente_fila = hoja_usuario.max_row + 1

    # Conversión e ingreso con factor 1000
    hoja_usuario.cell(row=siguiente_fila, column=1, value=datos_alimento["codigo"])
    hoja_usuario.cell(row=siguiente_fila, column=2, value=datos_alimento["alimento"])
    hoja_usuario.cell(row=siguiente_fila, column=3, value=int(round(datos_alimento["peso"] * FACTOR_ESCALA)))
    hoja_usuario.cell(row=siguiente_fila, column=4, value=int(round(datos_alimento["calorias"] * FACTOR_ESCALA)))
    hoja_usuario.cell(row=siguiente_fila, column=5, value=int(round(datos_alimento["proteinas"] * FACTOR_ESCALA)))
    hoja_usuario.cell(row=siguiente_fila, column=6, value=int(round(datos_alimento["grasas"] * FACTOR_ESCALA)))
    hoja_usuario.cell(row=siguiente_fila, column=7, value=int(round(datos_alimento["hidratos"] * FACTOR_ESCALA)))
    hoja_usuario.cell(row=siguiente_fila, column=8, value=int(round(datos_alimento["fibras"] * FACTOR_ESCALA)))


def procesar_comando_usuario(wb, entrada_usuario):
    """
    Función principal de procesamiento de comandos de entrada.
    """
    if "Comidas_Predeterminadas" not in wb.sheetnames:
        print("Error: La hoja 'Comidas_Predeterminadas' no existe.")
        return False, None

    hoja_predeterminadas = wb["Comidas_Predeterminadas"]

    resultado = buscar_comida_predeterminada(hoja_predeterminadas, entrada_usuario)

    if resultado is None:
        print(f"Error: Comando o alimento '{entrada_usuario}' no reconocido.")
        return False, None

    # Si la hoja de registro existe, registra convirtiendo al formato x1000
    if "Registro_Diario" in wb.sheetnames:
        hoja_usuario = wb["Registro_Diario"]
        registrar_alimento_en_usuario(hoja_usuario, resultado)
    
    print(f" Registrado con éxito: {resultado['alimento']}")
    print(f"   Valores Reales -> Peso: {resultado['peso']}g | Kcal: {resultado['calorias']} | Prot: {resultado['proteinas']}g")
    return True, resultado


# ==========================================
# INTERFAZ WEB (FLASK)
# ==========================================

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Buscador Nutricional</title>
    <style>
        body { font-family: Arial, sans-serif; max-width: 600px; margin: 40px auto; padding: 20px; background-color: #f4f6f8; }
        .card { background: white; padding: 25px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
        h1 { color: #2c3e50; text-align: center; margin-top: 0; }
        form { display: flex; gap: 10px; margin-bottom: 20px; }
        input[type="text"] { flex: 1; padding: 12px; border: 1px solid #ccc; border-radius: 5px; font-size: 16px; }
        button { padding: 12px 20px; background-color: #27ae60; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; font-weight: bold; }
        button:hover { background-color: #219150; }
        .result { background-color: #eef9f1; border-left: 5px solid #27ae60; padding: 15px; border-radius: 4px; margin-top: 20px; }
        .error { background-color: #fde8e8; border-left: 5px solid #e74c3c; color: #c0392b; padding: 15px; border-radius: 4px; margin-top: 20px; }
        table { width: 100%; border-collapse: collapse; margin-top: 10px; }
        th, td { text-align: left; padding: 8px; border-bottom: 1px solid #ddd; }
        th { background-color: #f8f9fa; }
    </style>
</head>
<body>
    <div class="card">
        <h1>🔍 Desglose Nutricional</h1>
        <form method="POST">
            <input type="text" name="comida" placeholder="Ingrese una comida..." value="{{ consulta if consulta else '' }}" required autofocus>
            <button type="submit">Buscar</button>
        </form>

        {% if error %}
            <div class="error">
                <strong>Error:</strong> {{ error }}
            </div>
        {% endif %}

        {% if resultado %}
            <div class="result">
                <h3>Resultados para: <em>{{ resultado.alimento }}</em> (Código: {{ resultado.codigo }})</h3>
                <table>
                    <tr><th>Componente</th><th>Cantidad</th></tr>
                    <tr><td><strong>Peso</strong></td><td>{{ resultado.peso }} g</td></tr>
                    <tr><td><strong>Calorías</strong></td><td>{{ resultado.calorias }} kcal</td></tr>
                    <tr><td><strong>Proteínas</strong></td><td>{{ resultado.proteinas }} g</td></tr>
                    <tr><td><strong>Grasas</strong></td><td>{{ resultado.grasas }} g</td></tr>
                    <tr><td><strong>Carbohidratos</strong></td><td>{{ resultado.hidratos }} g</td></tr>
                    <tr><td><strong>Fibra</strong></td><td>{{ resultado.fibras }} g</td></tr>
                </table>
            </div>
        {% endif %}
    </div>
</body>
</html>
"""

@app.route("/", methods=["GET", "POST"])
def home():
    resultado = None
    error = None
    consulta = ""

    if request.method == "POST":
        consulta = request.form.get("comida", "").strip()

        if not os.path.exists(EXCEL_FILE):
            error = f"No se encontró el archivo de datos '{EXCEL_FILE}'."
        else:
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
                exito, res = procesar_comando_usuario(wb, consulta)
                if exito:
                    resultado = res
                    # Guardamos si hubo cambios en la hoja usuario
                    if "Registro_Diario" in wb.sheetnames:
                        wb.save(EXCEL_FILE)
                else:
                    error = f"No se encontró el alimento o comando '{consulta}'."
            except Exception as e:
                error = f"Error al procesar el Excel: {str(e)}"

    return render_template_string(HTML_TEMPLATE, resultado=resultado, error=error, consulta=consulta)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
